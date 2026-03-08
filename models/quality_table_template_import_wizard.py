import base64
import io
import json

from odoo import api, fields, models
from odoo.exceptions import UserError


class QualityTableTemplateImportWizard(models.TransientModel):
    _name = "quality.table.template.import.wizard"
    _description = "质检表模板批量导入向导"

    file = fields.Binary("导入文件", required=True)
    file_name = fields.Char("文件名")
    source_format = fields.Selection(
        [("auto", "自动识别"), ("json", "JSON"), ("xlsx", "Excel(.xlsx)")],
        string="文件格式",
        default="auto",
        required=True,
    )
    import_mode = fields.Selection(
        [("upsert", "新增或更新"), ("create", "仅新增"), ("update", "仅更新")],
        string="导入模式",
        default="upsert",
        required=True,
    )
    dry_run = fields.Boolean("仅预演（不落库）", default=False)
    version_note = fields.Char("版本说明", default="批量导入")

    total_rows = fields.Integer("总行数", readonly=True)
    created_count = fields.Integer("新增数", readonly=True)
    updated_count = fields.Integer("更新数", readonly=True)
    skipped_count = fields.Integer("跳过数", readonly=True)
    error_count = fields.Integer("错误数", readonly=True)
    result_text = fields.Text("执行日志", readonly=True)

    HEADER_ALIASES = {
        "code": {"code", "模板编码", "编码", "template_code"},
        "name": {"name", "模板名称", "名称", "template_name"},
        "target_table_type": {"target_table_type", "目标类型", "表单类型", "type"},
        "sequence": {"sequence", "优先级"},
        "min_score": {"min_score", "最小命中分"},
        "keyword_rules": {"keyword_rules", "关键词规则", "命中规则"},
        "required_keywords": {"required_keywords", "必含关键词"},
        "title_regex": {"title_regex", "标题正则"},
        "field_patterns_json": {"field_patterns_json", "字段提取规则", "字段规则"},
        "default_values_json": {"default_values_json", "默认值", "默认值json"},
        "notes": {"notes", "备注"},
        "active": {"active", "启用"},
    }

    def _reopen_self(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": "模板批量导入",
            "res_model": "quality.table.template.import.wizard",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }

    @staticmethod
    def _clean_text(value):
        return str(value or "").replace("\x00", "").strip()

    @classmethod
    def _normalize_header(cls, header):
        text = cls._clean_text(header).lower()
        for field_name, aliases in cls.HEADER_ALIASES.items():
            if text in {alias.lower() for alias in aliases}:
                return field_name
        return text

    @staticmethod
    def _to_bool(value):
        text = str(value or "").strip().lower()
        if text in {"1", "true", "yes", "y", "是", "启用"}:
            return True
        if text in {"0", "false", "no", "n", "否", "停用"}:
            return False
        return bool(text)

    @staticmethod
    def _parse_int(value, default=0):
        if value in (None, ""):
            return default
        try:
            return int(float(str(value).strip()))
        except Exception:
            return default

    @staticmethod
    def _parse_target_type(value):
        text = str(value or "").strip().lower()
        mapping = {
            "7": "7",
            "桥施7": "7",
            "table7": "7",
            "13": "13",
            "桥施13": "13",
            "table13": "13",
            "other": "other",
            "通用": "other",
            "通用表单": "other",
        }
        return mapping.get(text, "other")

    @staticmethod
    def _ensure_json_text(value):
        if value in (None, ""):
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        text = str(value).strip()
        if not text:
            return ""
        try:
            loaded = json.loads(text)
            return json.dumps(loaded, ensure_ascii=False)
        except Exception:
            return text

    def _detect_format(self):
        self.ensure_one()
        if self.source_format != "auto":
            return self.source_format
        lower = (self.file_name or "").lower()
        if lower.endswith(".json"):
            return "json"
        if lower.endswith(".xlsx"):
            return "xlsx"
        raise UserError("无法自动识别文件格式，请手工选择 JSON 或 Excel(.xlsx)。")

    def _parse_json_rows(self, raw_bytes):
        text = ""
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                text = raw_bytes.decode(enc)
                break
            except Exception:
                continue
        if not text.strip():
            raise UserError("JSON文件内容为空。")
        try:
            payload = json.loads(text)
        except Exception as exc:
            raise UserError(f"JSON解析失败: {exc}") from exc

        if isinstance(payload, list):
            rows = payload
        elif isinstance(payload, dict):
            if isinstance(payload.get("templates"), list):
                rows = payload.get("templates")
            elif isinstance(payload.get("items"), list):
                rows = payload.get("items")
            elif isinstance(payload.get("data"), list):
                rows = payload.get("data")
            else:
                rows = [payload]
        else:
            raise UserError("JSON根节点必须是对象或数组。")

        parsed = []
        for row in rows:
            if isinstance(row, dict):
                parsed.append(row)
        return parsed

    def _parse_xlsx_rows(self, raw_bytes):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise UserError("当前环境缺少 openpyxl，无法导入 Excel。请改用 JSON 或安装 openpyxl。") from exc

        wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        headers = None
        rows = []
        for values in row_iter:
            if headers is None:
                headers = [self._normalize_header(v) for v in values]
                continue
            if not any(v not in (None, "") for v in values):
                continue
            item = {}
            for idx, val in enumerate(values):
                if idx >= len(headers):
                    continue
                key = headers[idx]
                if not key:
                    continue
                item[key] = val
            rows.append(item)
        return rows

    def _load_rows(self):
        self.ensure_one()
        if not self.file:
            raise UserError("请先上传导入文件。")
        raw_bytes = base64.b64decode(self.file)
        fmt = self._detect_format()
        if fmt == "json":
            return self._parse_json_rows(raw_bytes)
        if fmt == "xlsx":
            return self._parse_xlsx_rows(raw_bytes)
        raise UserError("不支持的文件格式。")

    def _normalize_row(self, row):
        normalized = {}
        for raw_key, value in (row or {}).items():
            key = self._normalize_header(raw_key)
            normalized[key] = value

        vals = {
            "code": self._clean_text(normalized.get("code")),
            "name": self._clean_text(normalized.get("name")),
            "target_table_type": self._parse_target_type(normalized.get("target_table_type")),
            "sequence": self._parse_int(normalized.get("sequence"), default=10),
            "min_score": self._parse_int(normalized.get("min_score"), default=1),
            "keyword_rules": self._clean_text(normalized.get("keyword_rules")),
            "required_keywords": self._clean_text(normalized.get("required_keywords")),
            "title_regex": self._clean_text(normalized.get("title_regex")),
            "field_patterns_json": self._ensure_json_text(normalized.get("field_patterns_json")),
            "default_values_json": self._ensure_json_text(normalized.get("default_values_json")),
            "notes": self._clean_text(normalized.get("notes")),
        }
        if "active" in normalized:
            vals["active"] = self._to_bool(normalized.get("active"))

        return vals

    def action_import(self):
        self.ensure_one()
        rows = self._load_rows()
        template_model = self.env["coordos.quality.table.template"]

        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0
        logs = []

        for idx, row in enumerate(rows, start=1):
            vals = self._normalize_row(row)
            code = vals.get("code")
            if not code:
                skipped_count += 1
                logs.append(f"第{idx}行: 跳过，缺少 code。")
                continue
            if not vals.get("name"):
                vals["name"] = code

            existing = template_model.search([("code", "=", code)], limit=1)
            try:
                if self.import_mode == "create" and existing:
                    skipped_count += 1
                    logs.append(f"第{idx}行[{code}]: 已存在，按模式跳过。")
                    continue
                if self.import_mode == "update" and not existing:
                    skipped_count += 1
                    logs.append(f"第{idx}行[{code}]: 不存在，按模式跳过。")
                    continue

                if self.dry_run:
                    if existing:
                        updated_count += 1
                        logs.append(f"第{idx}行[{code}]: 预演更新。")
                    else:
                        created_count += 1
                        logs.append(f"第{idx}行[{code}]: 预演新增。")
                    continue

                if existing:
                    existing.with_context(version_note=(self.version_note or "批量导入更新")).write(vals)
                    updated_count += 1
                    logs.append(f"第{idx}行[{code}]: 更新成功。")
                else:
                    template_model.with_context(version_note=(self.version_note or "批量导入创建")).create(vals)
                    created_count += 1
                    logs.append(f"第{idx}行[{code}]: 新增成功。")
            except Exception as exc:
                error_count += 1
                logs.append(f"第{idx}行[{code}]: 错误 - {exc}")

        self.total_rows = len(rows)
        self.created_count = created_count
        self.updated_count = updated_count
        self.skipped_count = skipped_count
        self.error_count = error_count
        self.result_text = "\n".join(logs[:500])
        return self._reopen_self()
